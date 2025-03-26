import pytesseract
from PIL import Image
import pyautogui
import time
import cv2
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Tesseract yolu belirtmeniz gerekebilir
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'  # Windows için örnek yol

# Paint Seed'ler için örnek diziler (değiştirilebilir)
de_pattern_1 = ["490", "148", "69", "704"]
de_rank_1 = ["16", "48", "66", "67", "96", "111", "117", "159", "259", "263", "273", "297", "308", "321", "324", "341",
             "347", "461", "482", "517", "530", "567", "587", "674", "695", "723", "764", "772", "781", "790", "792",
             "843", "880", "885", "904", "948", "990"]
de_rank_2 = ["09", "116", "134", "158", "168", "225", "338", "354", "356", "365", "370", "386", "406", "426", "433",
             "441", "483", "537", "542", "592", "607", "611", "651", "668", "673", "696", "730", "743", "820", "846",
             "856", "857", "870", "876", "878", "882", "898", "900", "925", "942", "946", "951", "953", "970", "998"]
de_purple = ["29", "31", "33", "43", "72", "85", "88", "99", "104", "105", "128", "133", "136", "139", "140", "146",
             "146", "156", "172", "174", "176", "189", "216", "217", "249", "265", "290", "293", "310", "310", "322",
             "339", "340", "343", "363", "395", "404", "411", "437", "449", "451", "453", "458", "463", "487", "496",
             "509", "532", "550", "572", "572", "574", "598", "599", "605", "606", "614", "621", "627", "631", "653",
             "666", "667", "672", "683", "707", "710", "717", "727", "734", "740", "750", "766", "778", "795", "800",
             "804", "806", "811", "815", "816", "817", "839", "842", "848", "849", "850", "862", "877", "891", "913",
             "926", "927", "944", "952", "969", "971", "989"]
de_gold = ["4", "6", "14", "24", "37", "74", "78", "79", "80", "87", "102", "103", "124", "132", "135", "144", "167",
           "177", "180", "181", "182", "184", "184", "184", "205", "243", "247", "248", "252", "256", "256", "264",
           "269", "270", "277", "289", "292", "301", "323", "325", "334", "360", "362", "367", "374", "382", "392",
           "403", "428", "432", "443", "446", "466", "491", "492", "495", "505", "511", "527", "549", "555", "555",
           "558", "564", "568", "568", "577", "594", "604", "608", "623", "624", "629", "637", "638", "645", "646",
           "646", "686", "692", "733", "735", "738", "746", "783", "798", "802", "803", "813", "837", "868", "868",
           "887", "903", "907", "911", "914", "921", "923", "945", "986", "994"]


# OCR ile Paint Seed verilerini alacak fonksiyon
def extract_text_from_image(image_path):
    # Görüntüyü yükle
    img = cv2.imread(image_path)

    # Görüntüyü gri tonlamaya çevir
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    # Görüntüdeki metni tespit et
    text = pytesseract.image_to_string(gray)

    # Paint Seed değerini metinden çıkar
    paint_seeds = []
    if "Paint Seed" in text:
        lines = text.splitlines()
        for line in lines:
            if "Paint Seed" in line:
                # Paint Seed bilgilerini çıkart
                paint_seed = line.split(":")[-1].strip()
                paint_seeds.append(paint_seed)
    return paint_seeds


def process_and_classify_paint_seeds(paint_seeds):
    # Paint Seed'leri sınıflandır
    classified_data = []

    for seed in paint_seeds:
        if seed in de_pattern_1:
            classified_data.append({"Paint Seed": seed, "Category": "Blue-Gem", "Color": "magenta"})
        elif seed in de_rank_1:
            classified_data.append({"Paint Seed": seed, "Category": "Rank 1", "Color": "magenta"})
        elif seed in de_rank_2:
            classified_data.append({"Paint Seed": seed, "Category": "Rank 2", "Color": "red"})
        elif seed in de_purple:
            classified_data.append({"Paint Seed": seed, "Category": "Purple", "Color": "yellow"})
        elif seed in de_gold:
            classified_data.append({"Paint Seed": seed, "Category": "Gold", "Color": "green"})
        else:
            classified_data.append({"Paint Seed": seed, "Category": "Unknown", "Color": "white"})

    return classified_data


def save_to_excel(classified_data):
    # Verileri pandas DataFrame'e dönüştür
    df = pd.DataFrame(classified_data)

    # Yeni CSV dosya adı oluşturma (zaman damgası kullanarak her seferinde yeni dosya adı)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"bluegem_finder_{timestamp}.xlsx"

    # Excel dosyasını oluştur
    wb = Workbook()
    ws = wb.active

    # DataFrame'i Excel'e yaz
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    # Renklerin stilini belirleme
    color_map = {
        "magenta": "FF00FF",
        "red": "FF0000",
        "yellow": "FFFF00",
        "green": "00FF00",
        "white": "FFFFFF"
    }

    # Satırları renklendir
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):  # Başlık satırını atla
        category_cell = row[1]
        color = category_cell.value.lower()

        # Belirli renklere göre hücre rengi ayarla
        fill = PatternFill(start_color=color_map.get(color, "FFFFFF"), end_color=color_map.get(color, "FFFFFF"),
                           fill_type="solid")
        for cell in row:
            cell.fill = fill

    # Excel dosyasına kaydet
    wb.save(file_name)
    print(f"Paint Seed verileri {file_name} dosyasına kaydedildi.")


# Ekran görüntüsünü alma fonksiyonu
def capture_full_screenshot():
    screenshot = pyautogui.screenshot()  # PyAutoGUI ile ekran görüntüsünü al
    screenshot_path = "screenshot.png"  # Kaydedilecek dosya yolu
    screenshot.save(screenshot_path)  # Görüntüyü kaydet
    return screenshot_path


# Ana monitörün çözünürlüğü (örnek çözünürlük)
screen_width = 1920  # Ana monitör genişliği
screen_height = 1080  # Ana monitör yüksekliği


# Sayfa kaydırma ve verileri işleme fonksiyonu
def scroll_page_on_main_monitor():
    # Paint Seed'lerini saklamak için bir liste oluştur
    paint_seed_list = []

    # Monitörde kaydırma işlemi
    for _ in range(10):  # 10 kez kaydırma yapacak
        # Ekran görüntüsünü al
        screenshot_path = capture_full_screenshot()

        # OCR ile Paint Seed verilerini çıkart
        paint_seeds = extract_text_from_image(screenshot_path)

        # Paint Seed'leri sınıflandır
        classified_data = process_and_classify_paint_seeds(paint_seeds)

        # Sınıflandırılmış verileri kaydet
        save_to_excel(classified_data)

        # Sayfa kaydırma
        pyautogui.scroll(-300)  # Aşağı kaydırma

        # Kısa süre bekleme
        time.sleep(2)


# Ana fonksiyon çalıştırma
scroll_page_on_main_monitor()
